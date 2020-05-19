from numpy import *
import json
from xlrd import open_workbook
from xlutils.copy import copy
import pandas as pd  #@xiaoyu
from openpyxl import load_workbook #@x
from openpyxl import Workbook #@x

def evaluation_F1(order, top_k, positive_item):
    epsilon = 0.1**10
    top_k_items = set(order[0: top_k])
    positive_item = set(positive_item)
    precision = 1.0 * len(top_k_items & positive_item) / max(len(top_k_items), epsilon)
    recall = 1.0 * len(top_k_items & positive_item) / max(len(positive_item), epsilon)
    F1 = 2.0 * precision * recall / max(precision + recall, epsilon)
    return F1

def evaluation_NDCG(order, top_k, positive_item):
    top_k_item = order[0: top_k]
    epsilon = 0.1**10
    DCG = 0
    iDCG = 0
    for i in range(top_k):
        if top_k_item[i] in positive_item:
            DCG += 1 / log2(i + 2)
    for i in range(min(len(positive_item), top_k)):
        iDCG += 1 / log2(i + 2)
    NDCG = 1.0 * DCG / max(iDCG, epsilon)
    return NDCG

def readdata(path):
    with open(path) as f:
        line = f.readline()
        data = json.loads(line)
    f.close()
    user_num = len(data)
    item_num = 0
    interactions = []
    for user in range(user_num):
        for item in data[user]:
            interactions.append((user, item))
            item_num = max(item, item_num)
    item_num += 1
    return(interactions, data, user_num, item_num)

#@x
def df2str(df):
    df_str = ''
    for i in range(df.shape[0]):
        df_list = df.iloc[[i], :].values.tolist()
        df_list2 = [str(i) for i in df_list]
        str_temp = ''.join(df_list2)
        df_str = df_str +str_temp+','
    return df_str

#@x
def save_df(df_list,path_excel,first_sheet):
    excelWriter = pd.ExcelWriter(path_excel, engine='openpyxl')

    if first_sheet is False:
        workbook = load_workbook(path_excel)
        excelWriter.book = workbook
        exist_sheets = workbook.get_sheet_names()
        for df in df_list:
            if df[1] in exist_sheets:
                workbook.remove_sheet(workbook.get_sheet_by_name(df[1]))
            df[0].to_excel(excel_writer=excelWriter, sheet_name=df[1],index = True)
            excelWriter.save()
    else:
        for df in df_list:
            df[0].to_excel(excel_writer=excelWriter, sheet_name=df[1], index=True)
            excelWriter.save()
    excelWriter.close()

#@x
def save_parameters(data,path_excel):
    wb = Workbook( )
    table = wb.active
    table.title = 'Parameters'
    ldata = []
    for x in data:
        t = [x[0]]
        for elements in x[1]:
            t.append(elements)
        ldata.append(t)
    for i, p in enumerate(ldata):
        for j, q in enumerate(p):
            table.cell(row = i+1, column = j+1).value = q
    wb.save(path_excel)